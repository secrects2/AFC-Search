"""Report service — generate Excel reports from database."""
from __future__ import annotations

import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from src.database import Database, SnapshotRow

LOGGER = logging.getLogger(__name__)


# Column headers for daily report
DAILY_COLUMNS = [
    ("商品名稱", "product_name"),
    ("品牌", "brand"),
    ("建議售價", "suggested_price"),
    ("平台", "platform"),
    ("賣場商品標題", "title"),
    ("賣家", "seller"),
    ("商品網址", "url"),
    ("最終價格", "final_price"),
    ("最終來源", "final_price_source"),
    ("最終信心度", "final_confidence"),
    ("最終狀態", "final_status"),
    ("直接抓價", "direct_price"),
    ("飛比抓價", "feebee_price"),
    ("LBJ抓價", "lbj_price"),
    ("人工價格", "manual_price"),
    ("決策原因", "decision_reason"),
    ("舊有抓價", "price"),
    ("價差", "price_diff"),
    ("是否疑似破價", "is_violation"),
    ("狀態", "candidate_status"),
    ("上次檢查時間", "checked_at"),
    ("第一次發現時間", "first_seen_at"),
    ("搜尋來源", "source_found_by"),
    ("截圖路徑", "screenshot_path"),
    ("錯誤訊息", "error_message"),
]


class ReportService:
    """Generate Excel reports from database snapshots."""

    def __init__(self, db: Database, project_root: Path) -> None:
        self.db = db
        self.output_dir = project_root / "output"
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_daily_report(self, date: str | None = None) -> Path:
        """Export daily_report.xlsx with all snapshots for a given date."""
        today = date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
        snapshots = self.db.get_snapshots(date=today)

        if not snapshots:
            # Fall back to latest snapshots if no data for today
            snapshots = self.db.get_snapshots(limit=500)

        path = self.output_dir / "daily_report.xlsx"
        self._write_xlsx(path, snapshots, f"每日報表 {today}")
        LOGGER.info("每日報表已產出：%s (%d 筆)", path, len(snapshots))
        return path

    def export_violations_report(self, date: str | None = None) -> Path:
        """Export suspected_violations.xlsx with only violation snapshots."""
        today = date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
        snapshots = self.db.get_snapshots(date=today, violation_only=True)

        if not snapshots:
            snapshots = self.db.get_snapshots(violation_only=True, limit=200)

        path = self.output_dir / "suspected_violations.xlsx"
        self._write_xlsx(path, snapshots, f"疑似破價 {today}")
        LOGGER.info("破價報表已產出：%s (%d 筆)", path, len(snapshots))
        return path

    def _write_xlsx(
        self, path: Path, snapshots: list[SnapshotRow], sheet_title: str
    ) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            LOGGER.warning("openpyxl 未安裝，改用 CSV 輸出")
            self._write_csv_fallback(path.with_suffix(".csv"), snapshots)
            return

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title[:31]  # Excel sheet name limit

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2B579A")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        for col_idx, (label, _) in enumerate(DAILY_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Write data rows
        violation_fill = PatternFill("solid", fgColor="FFF2CC")
        for row_idx, snap in enumerate(snapshots, start=2):
            # Parse raw_data once per row
            try:
                raw_data_dict = json.loads(snap.raw_data) if snap.raw_data else {}
            except Exception:
                raw_data_dict = {}
                
            for col_idx, (_, attr) in enumerate(DAILY_COLUMNS, start=1):
                if attr == "price_source":
                    value = raw_data_dict.get("price_source", "dom")
                elif attr == "direct_price":
                    value = raw_data_dict.get("all_prices", {}).get("direct_html", "")
                elif attr == "feebee_price":
                    value = raw_data_dict.get("all_prices", {}).get("feebee", "")
                elif attr == "lbj_price":
                    value = raw_data_dict.get("all_prices", {}).get("lbj", "")
                elif attr == "manual_price":
                    value = raw_data_dict.get("all_prices", {}).get("manual", "")
                elif attr == "is_violation":
                    value = "是" if snap.is_violation else ""
                elif attr == "candidate_status":
                    status_map = {
                        "normal": "正常",
                        "suspected_violation": "監控中",
                        "price_unknown": "未抓到價格",
                        "error": "錯誤",
                        "blocked": "遭到阻擋",
                        "takedown_notified": "取消監控",
                        "source_dead": "網址失效",
                    }
                    value = status_map.get(getattr(snap, attr, ""), getattr(snap, attr, ""))
                elif attr == "final_status":
                    final_status_map = {
                        "verified_price": "高可信",
                        "likely_price": "中可信",
                        "needs_review": "需審核",
                        "price_unknown": "未抓到價格",
                        "suspected_violation": "疑似破價",
                        "verified_violation": "確認破價",
                    }
                    val = getattr(snap, "final_status", "")
                    value = final_status_map.get(val, val)
                elif attr in ("price", "suggested_price", "price_diff", "final_price"):
                    raw_val = getattr(snap, attr, None)
                    if attr == "price_diff" and hasattr(snap, "final_price") and snap.final_price and snap.suggested_price:
                         raw_val = snap.suggested_price - snap.final_price
                    value = int(raw_val) if raw_val is not None else ""
                else:
                    value = getattr(snap, attr, "")
                    
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if snap.is_violation or getattr(snap, "final_status", "") in ("suspected_violation", "verified_violation"):
                    cell.fill = violation_fill

        # Auto-adjust column widths
        for col_idx, (label, _) in enumerate(DAILY_COLUMNS, start=1):
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = max(
                len(label) * 2, 12
            )

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(path)

    def _write_csv_fallback(self, path: Path, snapshots: list[SnapshotRow]) -> None:
        import csv
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([label for label, _ in DAILY_COLUMNS])
            for snap in snapshots:
                try:
                    raw_data_dict = json.loads(snap.raw_data) if snap.raw_data else {}
                except Exception:
                    raw_data_dict = {}
                    
                row = []
                for _, attr in DAILY_COLUMNS:
                    if attr == "price_source":
                        value = raw_data_dict.get("price_source", "dom")
                    elif attr == "direct_price":
                        value = raw_data_dict.get("all_prices", {}).get("direct_html", "")
                    elif attr == "feebee_price":
                        value = raw_data_dict.get("all_prices", {}).get("feebee", "")
                    elif attr == "lbj_price":
                        value = raw_data_dict.get("all_prices", {}).get("lbj", "")
                    elif attr == "manual_price":
                        value = raw_data_dict.get("all_prices", {}).get("manual", "")
                    elif attr == "is_violation":
                        value = "是" if snap.is_violation else ""
                    elif attr == "candidate_status":
                        status_map = {
                            "normal": "正常",
                            "suspected_violation": "監控中",
                            "price_unknown": "未抓到價格",
                            "error": "錯誤",
                            "blocked": "遭到阻擋",
                            "takedown_notified": "取消監控",
                            "source_dead": "網址失效",
                        }
                        value = status_map.get(getattr(snap, attr, ""), getattr(snap, attr, ""))
                    elif attr == "final_status":
                        final_status_map = {
                            "verified_price": "高可信",
                            "likely_price": "中可信",
                            "needs_review": "需審核",
                            "price_unknown": "未抓到價格",
                            "suspected_violation": "疑似破價",
                            "verified_violation": "確認破價",
                        }
                        val = getattr(snap, "final_status", "")
                        value = final_status_map.get(val, val)
                    elif attr in ("price", "suggested_price", "price_diff", "final_price"):
                        raw_val = getattr(snap, attr, None)
                        if attr == "price_diff" and hasattr(snap, "final_price") and snap.final_price and snap.suggested_price:
                            raw_val = snap.suggested_price - snap.final_price
                        value = int(raw_val) if raw_val is not None else ""
                    else:
                        value = getattr(snap, attr, "")
                    row.append(value)
                writer.writerow(row)
